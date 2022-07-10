import requests
import nltk
from bs4 import BeautifulSoup
import pandas as pd
from nltk.tokenize import sent_tokenize, word_tokenize
import openpyxl
nltk.download('punkt')

headers = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.12; rv:55.0) Gecko/20100101 Firefox/55.0',
}

data = pd.read_csv('Input.csv')
names = data.iloc[:170,0].values
links = data.iloc[:170,1].values

def headline():
  h = soup.find('h1', {'class':"entry-title"})
  strhead = str(h)
  return BeautifulSoup(strhead, "html.parser").get_text()
def bodytxt():
  b = soup.find('div', {'class':"td-post-content"})
  strbody = str(b)
  return BeautifulSoup(strbody, "html.parser").get_text()
stop = open('StopWords_Generic.txt','r')
pos_neg = pd.read_csv('check.csv')
wb_obj = openpyxl.load_workbook('Output.xlsx')
sheet = wb_obj.active
complex_word =0
positive_word =0
negative_word =0
word_count =0
word_un_count =0
personal_pronouns =0
word_ch =0
syllable_count =0 
syllable = 0
vowel = ['a','e','i','o','u']
ending =["es","ed"]
symbols = [';', ':', '!', "*", "?",',',"’",'”','(',')','.','“']
pronouns = ["I","we","my","ours" ,"us","We","My","Ours" ,"Us","our","Our"]
stop_word = word_tokenize(stop.read())
pos = []
neg = []
for x in range(len(pos_neg)):
  if pos_neg.iloc[x,7]>0:
    neg.append(pos_neg.iloc[x,0])
  elif pos_neg.iloc[x,8]>0:
    pos.append(pos_neg.iloc[x,0])

for x in range(170):
  url = links[x]
  page=requests.get(url,headers =headers)
  soup = BeautifulSoup(page.content, 'html.parser')
  f= open(str(int(names[x])),"w+")
  f.write(headline())
  f.write(bodytxt())
  f.close()


n =[]
for x in range(170):
  f = open(str(int(names[x])),"r")
  n.append(f.read())
  f.close()


for x in range(len(n)):
  sen = sent_tokenize(n[x])
  no_sen = len(sen)

  for z in sen:
    word_tokens = word_tokenize(z)
    cleared = [w for w in word_tokens if not w.lower() in symbols]
    word_un_count += len(cleared)
    wordcleared =[]


    for y in cleared:

      if y in pronouns: 
        personal_pronouns +=1 
    
      if y.upper() not in stop_word:
        wordcleared.append(y)





    print(wordcleared)
    word_count += len(wordcleared)
    for g in wordcleared:
        if g in pos:
          positive_word+=1
        elif g in neg:
          negative_word+=1
        
        syllable = 0
        word_ch+= len(g)
        for y in g:
          if y.lower() in vowel:
            syllable+=1
        if y[len(y)-2:len(y)] in ending:
          syllable-=1
        if syllable >2:
          complex_word+=1
        syllable_count += syllable

  syllable_per = syllable_count/word_count
  avg_word_sen = word_count/no_sen
  avg_no_word_sen = word_un_count/no_sen
  avg_word_len = word_ch/word_count
  polarity_score = (positive_word-negative_word)/((positive_word+negative_word)+0.000001)
  subjectivity_score = (positive_word+negative_word)/(word_un_count +0.000001)
  percent_complex = complex_word/word_count
  fog_index = 0.4*(percent_complex+avg_word_sen)
  sheet.cell(row = x+2,column =3).value = positive_word
  sheet.cell(row = x+2,column =4).value = negative_word
  sheet.cell(row = x+2,column =5).value = polarity_score
  sheet.cell(row = x+2,column =6).value = subjectivity_score
  sheet.cell(row = x+2,column =7).value = avg_word_sen
  sheet.cell(row = x+2,column =8).value = percent_complex
  sheet.cell(row = x+2,column =9).value = fog_index
  sheet.cell(row = x+2,column =10).value = avg_no_word_sen
  sheet.cell(row = x+2,column =11).value = complex_word
  sheet.cell(row = x+2,column =12).value = word_count
  sheet.cell(row = x+2,column =13).value = syllable_per
  sheet.cell(row = x+2,column =14).value = personal_pronouns
  sheet.cell(row = x+2,column =15).value = avg_word_len

wb_obj.save('Output.xlsx')
